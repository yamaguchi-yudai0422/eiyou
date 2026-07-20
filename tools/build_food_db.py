"""日本食品標準成分表(八訂) の Excel から アプリ用 CSV を生成する。

使い方:
    python tools/build_food_db.py <成分表.xlsx> [-o assets/data/foods.csv]

成分表の Excel は文部科学省のサイトから入手する:
    https://www.mext.go.jp/a_menu/syokuhinseibun/mext_00001.html
    「第2章（データ）」の xlsx

このスクリプトが出力するのは栄養価のみ。読み仮名・比重・1個あたりの重量は
成分表に収載されていないため、assets/data/food_overrides.csv で別に管理する。
栄養価（公式データ）と入力補助のための手入力データを混ぜないための分離。
"""

import argparse
import csv
import sys

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl が必要です:  pip install openpyxl")


# 成分識別子が並ぶ行と、データが始まる行。
# 成分表の版が変わってレイアウトがずれた場合はここを直す。
ID_ROW = 12
DATA_START_ROW = 13

# アプリ側の栄養素名 -> 成分表の成分識別子
# 並び順は lib/core/nutrition/nutrient.dart の NutrientDef.all と一致させること。
NUTRIENT_IDS = [
    ("energy", "ENERC_KCAL"),
    ("protein", "PROT-"),
    ("fat", "FAT-"),
    ("carbohydrate", "CHOCDF-"),
    ("fiber", "FIB-"),
    ("salt", "NACL_EQ"),
    ("calcium", "CA"),
    ("iron", "FE"),
    ("magnesium", "MG"),
    ("potassium", "K"),
    ("zinc", "ZN"),
    ("vitaminA", "VITA_RAE"),
    ("vitaminD", "VITD"),
    ("vitaminE", "TOCPHA"),
    ("vitaminK", "VITK"),
    ("vitaminB1", "THIA"),
    ("vitaminB2", "RIBF"),
    # ナイアシンは「ナイアシン当量(NE)」を採る。食事摂取基準が NE 基準のため。
    ("niacin", "NE"),
    ("vitaminB6", "VITB6A"),
    ("vitaminB12", "VITB12"),
    ("folate", "FOL"),
    ("vitaminC", "VITC"),
]

# 食品群番号 -> 食品群名
FOOD_GROUPS = {
    "01": "穀類",
    "02": "いも類",
    "03": "砂糖類",
    "04": "豆類",
    "05": "種実類",
    "06": "野菜類",
    "07": "果実類",
    "08": "きのこ類",
    "09": "藻類",
    "10": "魚介類",
    "11": "肉類",
    "12": "卵類",
    "13": "乳類",
    "14": "油脂類",
    "15": "菓子類",
    "16": "し好飲料類",
    "17": "調味料",
    "18": "調理済み流通food",
}
FOOD_GROUPS["18"] = "調理済み食品"

CSV_HEADER = [
    "food_code",
    "name",
    "category",
] + [name for name, _ in NUTRIENT_IDS]


def parse_value(raw):
    """成分表の値を float にする。

    成分表の表記:
        None / ""  未収載
        "-"        未測定
        "Tr"       微量（Trace）。0 として扱う。
        "(12.3)"   推計値。括弧を外して採用する。
        "(Tr)"     推計値の微量。0。
    栄養計算では未測定と微量をどちらも 0 として扱う。過大評価を避けるため。
    """
    if raw is None:
        return 0.0
    if isinstance(raw, (int, float)):
        return float(raw)

    s = str(raw).strip()
    s = s.replace("（", "(").replace("）", ")")
    if s.startswith("(") and s.endswith(")"):
        s = s[1:-1].strip()
    if s in ("", "-", "Tr", "tr", "*", "未測定"):
        return 0.0

    s = s.replace(",", "")
    try:
        return float(s)
    except ValueError:
        return 0.0


def build_column_index(id_row):
    """成分識別子 -> 列番号 の対応を作る。

    識別子には末尾に空白が入っているもの（'VITK '）があるので strip して比較する。
    """
    index = {}
    for col, value in enumerate(id_row):
        if value is None:
            continue
        key = str(value).strip()
        if key and key not in index:
            index[key] = col
    return index


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("xlsx", help="成分表 第2章（データ）の xlsx")
    parser.add_argument("-o", "--out", default="assets/data/foods.csv")
    args = parser.parse_args()

    wb = openpyxl.load_workbook(args.xlsx, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]  # 本表

    rows = list(ws.iter_rows(values_only=True))
    id_index = build_column_index(rows[ID_ROW - 1])

    missing = [cid for _, cid in NUTRIENT_IDS if cid not in id_index]
    if missing:
        sys.exit(
            "成分識別子が見つかりません: %s\n"
            "成分表の版が変わってレイアウトが変わった可能性があります。"
            % ", ".join(missing)
        )

    records = []
    skipped = 0
    for row in rows[DATA_START_ROW - 1:]:
        if not row or len(row) < 4:
            continue
        group_no = row[0]
        food_code = row[1]
        name = row[3]

        # 食品番号と食品名が揃っている行だけを採る（小計行・空行を除く）。
        if not food_code or not name:
            skipped += 1
            continue

        group_key = str(group_no).strip().zfill(2) if group_no else ""
        category = FOOD_GROUPS.get(group_key, "その他")

        # 成分表の食品名は全角空白で階層を区切っている。そのまま保持する。
        clean_name = str(name).replace("　", " ").strip()
        # アプリ側の CSV パーサは単純な split(',') なので、
        # 食品名にカンマが入ると列がずれる。全角カンマへ置き換えて防ぐ。
        clean_name = clean_name.replace(",", "、")

        values = [
            parse_value(row[id_index[cid]]) if id_index[cid] < len(row) else 0.0
            for _, cid in NUTRIENT_IDS
        ]

        records.append([str(food_code).strip(), clean_name, category] + values)

    with open(args.out, "w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        w.writerow(CSV_HEADER)
        w.writerows(records)

    print(f"{len(records)} 件を書き出しました -> {args.out}")
    print(f"（見出し・空行として除外: {skipped} 行）")

    by_group = {}
    for r in records:
        by_group[r[2]] = by_group.get(r[2], 0) + 1
    for group, count in sorted(by_group.items(), key=lambda kv: -kv[1]):
        print(f"  {group}: {count}")


if __name__ == "__main__":
    main()
